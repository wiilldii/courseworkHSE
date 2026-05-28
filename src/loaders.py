"""
Raw-data loaders.  Each loader returns a pandas DataFrame with clean column
names; no calculations are done here.
"""

import logging
import pandas as pd
import openpyxl

from src.config import (
    DATA_RAW,
    MISSING_MARKERS,
    JAPAN_AREA_CODE,
    TOKYO_MA_AREA_CODES,
    OSAKA_KANSAI_AREA_CODES,
    ESTAT_COL_HIERARCHY,
    ESTAT_COL_INDUSTRY,
    ESTAT_COL_MGMT_NO,
    ESTAT_DATA_START,
    MIC_TABLE_SPECS,
    WB_YEARS,
    WB_COUNTRY_CODE,
    JPO_PATENT_MANUAL_FILENAME,
    JPO_LEGACY_PATENT_MANUAL_FILENAME,
)
from src.utils import to_numeric, is_missing

log = logging.getLogger("coursework")

# ── helpers ───────────────────────────────────────────────────────────────────

def _is_prefecture_level(area_code: str) -> bool:
    """True for 'XXYYYY_Name' where YYY == '000' (prefecture-level rows)."""
    if not isinstance(area_code, str):
        return False
    code_part = area_code.split("_")[0]
    return len(code_part) == 5 and code_part[2:] == "000"


def _read_xlsx_rows(filename: str, sheet: str | None = None, skip: int = 0):
    """Yield row tuples from an Excel file using openpyxl (memory-efficient)."""
    path = DATA_RAW / filename
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= skip:
            yield row
    wb.close()


# ── Economic Census b1_006_1e ─────────────────────────────────────────────────

def load_economic_census() -> pd.DataFrame:
    """
    Load b1_006_1e.xlsx.
    Returns prefecture-level rows (+ Japan total) with:
      area_code, area_label, hierarchy, div_code, industry_label,
      establishments, employment
    """
    filename = "b1_006_1e.xlsx"
    log.info(f"Loading {filename}")

    TARGET_AREAS = set([JAPAN_AREA_CODE] + TOKYO_MA_AREA_CODES + OSAKA_KANSAI_AREA_CODES)

    rows = []
    skipped = 0
    for raw in _read_xlsx_rows(filename, skip=9):  # row 10 onward (0-indexed: skip 9 header rows)
        area = raw[1]
        if area not in TARGET_AREAS:
            skipped += 1
            continue

        div_code = raw[3]
        industry_label = raw[4]
        establishments = to_numeric(raw[5])
        employment = to_numeric(raw[6])

        rows.append({
            "area_code": area,
            "hierarchy": raw[2],
            "div_code": div_code,
            "industry_label": industry_label,
            "establishments": establishments,
            "employment": employment,
        })

    df = pd.DataFrame(rows)
    log.info(f"  {filename}: {len(df)} rows kept, {skipped:,} municipality rows skipped")
    return df


# ── METI 第1表: 2-digit manufacturing ─────────────────────────────────────────

def load_meti_2digit() -> pd.DataFrame:
    """
    Load r03_seizochiiki_tkh.xlsx, sheet 第1表.
    Returns rows for target prefectures + Japan total.
    Monetary values already in million yen (百万円).

    Columns: pref_code, pref_name, industry_code_2d, industry_name_jp,
             establishments, employment, raw_materials_mn_yen,
             shipments_mn_yen, value_added_mn_yen
    """
    filename = "r03_seizochiiki_tkh.xlsx"
    sheet = "第1表"
    log.info(f"Loading {filename} sheet {sheet}")

    from src.config import TOKYO_MA_PREF_CODES, OSAKA_KANSAI_PREF_CODES

    target_codes = set([JAPAN_AREA_CODE.split("_")[0][:2]] +
                       TOKYO_MA_PREF_CODES + OSAKA_KANSAI_PREF_CODES)
    # '00' = Japan total in METI
    target_codes = {"00"} | set(TOKYO_MA_PREF_CODES) | set(OSAKA_KANSAI_PREF_CODES)

    rows = []
    for raw in _read_xlsx_rows(filename, sheet=sheet, skip=9):  # data from row 10
        if raw[1] is None:
            continue
        pref_code = str(raw[5]).strip().zfill(2) if raw[5] is not None else None
        if pref_code not in target_codes:
            continue

        rows.append({
            "pref_code": pref_code,
            "pref_name": raw[6],
            "industry_code_2d": str(raw[3]).strip().zfill(2) if raw[3] is not None else None,
            "industry_name_jp": raw[4],
            "establishments": to_numeric(raw[7]),
            "employment": to_numeric(raw[8]),
            # col 9 = personnel costs (百万円) — not used for LQ
            "shipments_mn_yen": to_numeric(raw[11]),   # 製造品出荷額等
            "value_added_mn_yen": to_numeric(raw[12]),  # 付加価値額
        })

    df = pd.DataFrame(rows)
    log.info(f"  {sheet}: {len(df)} rows")
    return df


# ── METI 第7表: 4-digit manufacturing ─────────────────────────────────────────

def load_meti_4digit() -> pd.DataFrame:
    """
    Load r03_seizochiiki_tkh.xlsx, sheet 第7表.
    Values are in 万円 → converted to million yen (÷ 100).
    Loads ALL prefectures so Japan totals can be derived by summation.

    Columns: pref_code, pref_name, industry_code_4d, industry_name_jp,
             establishments, employment, shipments_mn_yen, value_added_mn_yen
    """
    filename = "r03_seizochiiki_tkh.xlsx"
    sheet = "第7表"
    log.info(f"Loading {filename} sheet {sheet}")

    # Load all prefectures (needed to aggregate Japan totals at 4-digit level)
    target_codes = None  # load everything

    rows = []
    for raw in _read_xlsx_rows(filename, sheet=sheet, skip=9):
        if raw[1] is None:
            continue
        pref_code = str(raw[3]).strip().zfill(2) if raw[3] is not None else None
        if pref_code is None:
            continue
        # Skip the Japan-aggregate row (pref_code "00") since we'll derive Japan
        # totals by summing prefecture rows below
        if pref_code == "00":
            continue

        ind_code = str(raw[5]).strip() if raw[5] is not None else None

        # Convert 万円 → million yen (÷ 100)
        def conv(v):
            n = to_numeric(v)
            return n / 100.0 if n is not None else None

        rows.append({
            "pref_code": pref_code,
            "pref_name": raw[4],
            "industry_code_4d": ind_code,
            "industry_name_jp": raw[6],
            "establishments": to_numeric(raw[7]),
            "employment": to_numeric(raw[8]),
            "shipments_mn_yen": conv(raw[11]),
            "value_added_mn_yen": conv(raw[12]),
        })

    df = pd.DataFrame(rows)
    log.info(f"  {sheet}: {len(df)} rows (monetary values converted to million yen)")
    return df


# ── MIC R&D summary tables (ka101, ka102) ─────────────────────────────────────

def inspect_mic_workbook(filename: str, max_rows: int = 20,
                         max_cols: int = 20, print_output: bool = True) -> dict:
    """
    Inspect an MIC/eStat Excel workbook without applying parsing assumptions.
    Returns sheet names plus the first max_rows x max_cols cells from the active
    sheet; optionally prints the same diagnostics for source verification.
    """
    path = DATA_RAW / filename
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=max_rows,
                            max_col=max_cols, values_only=True):
        rows.append(list(row))

    result = {
        "filename": filename,
        "sheet_names": wb.sheetnames,
        "active_sheet": ws.title,
        "first_rows": rows,
    }
    wb.close()

    if print_output:
        print(f"===== {filename} =====")
        print(f"sheet_names: {result['sheet_names']}")
        for idx, row in enumerate(rows, start=1):
            values = ["" if v is None else str(v).replace("\n", " / ") for v in row]
            print(f"{idx:02d}: " + " | ".join(values))

    return result


def _load_estat_table(filename: str, data_col_names: dict[int, str],
                      hierarchy_filter=None) -> pd.DataFrame:
    """
    Generic loader for eStat-format MIC tables (ka1xx / ka2xx).
    data_col_names: {col_offset_from_data_start: output_name}
    hierarchy_filter: if set, keep only rows where hierarchy == this value
    """
    log.info(f"Loading {filename}")
    spec = MIC_TABLE_SPECS.get(filename, {})
    sheet = spec.get("sheet")
    data_start_row = spec.get("data_start_row", 13)
    log.info(
        f"  MIC structure verified: sheet={sheet or 'active'}, "
        f"header_row={spec.get('header_row')}, unit_row={spec.get('unit_row')}, "
        f"data_start_row={data_start_row}"
    )

    rows = []
    for raw in _read_xlsx_rows(filename, sheet=sheet, skip=data_start_row - 1):
        if raw[ESTAT_COL_HIERARCHY] is None:
            continue
        if raw[ESTAT_COL_INDUSTRY] is None:
            continue
        hier = raw[ESTAT_COL_HIERARCHY]
        item = str(raw[ESTAT_COL_INDUSTRY]).strip()

        # Guard against accidentally reading Japanese header rows as data.
        if str(hier).strip() == "階層" or item == "項目":
            continue

        try:
            hier_int = int(hier)
        except (ValueError, TypeError):
            continue

        if hierarchy_filter is not None:
            if hier_int != hierarchy_filter:
                continue

        record = {
            "hierarchy": hier_int,
            "industry_jp": item,
            "row_no": raw[ESTAT_COL_MGMT_NO],
        }
        for offset, col_name in data_col_names.items():
            raw_val = raw[ESTAT_DATA_START + offset] if ESTAT_DATA_START + offset < len(raw) else None
            record[col_name] = to_numeric(raw_val) if not is_missing(raw_val) else None

        rows.append(record)

    df = pd.DataFrame(rows)
    log.info(f"  {filename}: {len(df)} rows")
    return df


def load_rd_personnel_summary() -> pd.DataFrame:
    """2024ka101: R&D personnel by research performer (national summary)."""
    return _load_estat_table(
        "2024ka101.xlsx",
        data_col_names={
            0: "n_research_performers",
            1: "total_rd_personnel",
            2: "researchers",
            3: "research_assistants",
            4: "technicians",
            5: "other_rd_staff",
        },
        hierarchy_filter=None,
    )


def load_rd_expenditure_summary() -> pd.DataFrame:
    """2024ka102: R&D expenditure by research performer (national summary)."""
    return _load_estat_table(
        "2024ka102.xlsx",
        data_col_names={
            0: "n_research_performers",
            1: "rd_expenditure_mn_yen",
            2: "personnel_costs_mn_yen",
            3: "raw_materials_mn_yen",
            4: "tangible_fixed_assets_mn_yen",
            5: "intangible_fixed_assets_mn_yen",
            6: "lease_fees_mn_yen",
            7: "other_expenses_mn_yen",
            8: "received_research_funds_mn_yen",
            9: "external_research_expenditure_mn_yen",
        },
        hierarchy_filter=None,
    )


def load_enterprise_rd_by_industry() -> pd.DataFrame:
    """2024ka201: Enterprise R&D by industry (national context)."""
    return _load_estat_table(
        "2024ka201.xlsx",
        data_col_names={
            0: "n_enterprises",
            1: "sample_enterprises",
            2: "total_employees",
            3: "total_sales_100mn_yen",  # source unit: 億円
            4: "n_rd_enterprises",
            5: "rd_enterprise_share_pct",
            6: "n_inhouse_rd_enterprises",
            7: "rd_enterprise_employees",
            8: "rd_enterprise_sales_100mn_yen",  # source unit: 億円
            9: "researchers",
            10: "researchers_total",
            17: "rd_expenditure_mn_yen",  # 社内使用研究費 (100万円)
            34: "external_research_expenditure_mn_yen",
            38: "rd_expenditure_to_sales_pct",
        },
        hierarchy_filter=None,
    )


# ── Technology exchange tables (ka210, ka211, ka212) ─────────────────────────

def load_tech_exports() -> pd.DataFrame:
    """2024ka210: Technology export receipts by industry."""
    return _load_estat_table(
        "2024ka210.xlsx",
        data_col_names={
            0: "n_exporting_firms",
            1: "n_exporting_firms_with_inhouse_rd",
            2: "total_sales_100mn_yen",  # source unit: 億円
            3: "inhouse_rd_expenditure_mn_yen",
            4: "tech_export_receipts_mn_yen",  # 対価受取額 (100万円)
            5: "tech_export_receipts_parent_subsidiary_mn_yen",
        },
        hierarchy_filter=None,
    )


def load_tech_imports() -> pd.DataFrame:
    """2024ka211: Technology import payments by industry."""
    return _load_estat_table(
        "2024ka211.xlsx",
        data_col_names={
            0: "n_importing_firms",
            1: "n_importing_firms_with_inhouse_rd",
            2: "total_sales_100mn_yen",  # source unit: 億円
            3: "inhouse_rd_expenditure_mn_yen",
            4: "tech_import_payments_mn_yen",  # 対価支払額 (100万円)
            5: "tech_import_payments_parent_subsidiary_mn_yen",
        },
        hierarchy_filter=None,
    )


def load_tech_exchange_by_region() -> pd.DataFrame:
    """2024ka212: Technology exchange by industry and world region."""
    return _load_estat_table(
        "2024ka212.xlsx",
        data_col_names={
            0: "tech_receipts_mn_yen",
            1: "tech_receipts_east_southeast_asia_mn_yen",
            2: "tech_receipts_west_asia_mn_yen",
            3: "tech_receipts_north_america_mn_yen",
            4: "tech_receipts_south_america_mn_yen",
            5: "tech_receipts_europe_mn_yen",
            6: "tech_receipts_other_mn_yen",
            7: "tech_payments_mn_yen",
            8: "tech_payments_north_america_mn_yen",
            9: "tech_payments_europe_mn_yen",
            10: "tech_payments_other_mn_yen",
        },
        hierarchy_filter=None,
    )


# ── JPO patent data ───────────────────────────────────────────────────────────

def load_jpo_patent_applications() -> pd.DataFrame | None:
    """
    Load manual JPO prefecture patent application data if present.
    Preferred file: jpo_patent_applications_prefecture_manual.csv.
    Legacy file name is accepted for backward compatibility.
    """
    preferred = DATA_RAW / JPO_PATENT_MANUAL_FILENAME
    legacy = DATA_RAW / JPO_LEGACY_PATENT_MANUAL_FILENAME
    path = preferred if preferred.exists() else legacy if legacy.exists() else None
    if path is None:
        log.warning(
            "  JPO patent manual CSV not found; patent component skipped from HTP index"
        )
        return None

    df = pd.read_csv(path, dtype={"prefecture_code": str}, encoding="utf-8-sig")
    required = {
        "prefecture_code",
        "prefecture_jp",
        "prefecture_en",
        "patent_applications_2024",
    }
    missing = sorted(required - set(df.columns))
    if missing:
        raise ValueError(f"{path.name} missing required columns: {missing}")

    df["prefecture_code"] = df["prefecture_code"].astype(str).str.zfill(2)
    for col in [c for c in df.columns if c.startswith("patent_applications_")]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    log.info(f"  Loaded JPO patent applications: {path.name} ({len(df)} rows)")
    return df


# ── World Bank high-tech exports ──────────────────────────────────────────────

def load_worldbank_hightech() -> dict[str, pd.DataFrame]:
    """
    Load World Bank high-tech export CSVs.
    Returns dict with keys 'usd' and 'share_manufactured'.
    Each DataFrame has columns: year, value.
    """
    results = {}

    files = {
        "usd": "API_TX.VAL.TECH.CD_DS2_en_csv_v2_6018.csv",
        "share_manufactured": "API_TX.VAL.TECH.MF.ZS_DS2_en_csv_v2_1859.csv",
    }

    for key, filename in files.items():
        path = DATA_RAW / filename
        log.info(f"Loading {filename}")
        df_wide = pd.read_csv(path, skiprows=4, encoding="utf-8-sig")
        df_jpn = df_wide[df_wide["Country Code"] == WB_COUNTRY_CODE].copy()

        if df_jpn.empty:
            log.warning(f"  No JPN row found in {filename}")
            results[key] = pd.DataFrame(columns=["year", "value"])
            continue

        year_cols = [str(y) for y in WB_YEARS if str(y) in df_wide.columns]
        row = df_jpn.iloc[0]

        records = []
        for yr in year_cols:
            val = row[yr]
            if is_missing(val) or (isinstance(val, str) and val.strip() == ""):
                records.append({"year": int(yr), "value": None})
            else:
                try:
                    records.append({"year": int(yr), "value": float(val)})
                except (ValueError, TypeError):
                    records.append({"year": int(yr), "value": None})

        df_long = pd.DataFrame(records)
        log.info(f"  {filename}: {len(df_long)} year rows for JPN")
        results[key] = df_long

    return results
